# =========================
# FOOTBALL SIGNAL BOT
# =========================

import os
import asyncio
import logging
from datetime import datetime

import aiohttp
from aiogram import Bot, Dispatcher
from aiogram.filters import Command
from aiogram.types import Message
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

load_dotenv()
logging.basicConfig(level=logging.INFO)

BOT_TOKEN = os.getenv("BOT_TOKEN")
ADMIN_IDS = [int(x) for x in os.getenv("ADMIN_IDS", "").split(",")]
CHANNEL_ID = int(os.getenv("CHANNEL_ID"))
API_KEY = os.getenv("APIFOOTBALL_KEY")

API_BASE = "https://v3.football.api-sports.io"
STAKE = float(os.getenv("STAKE_EUR", 7))

# =========================
# SIMPLE API CLIENT
# =========================

async def api_get(endpoint, params=None):
    headers = {"x-apisports-key": API_KEY}
    async with aiohttp.ClientSession() as session:
        async with session.get(f"{API_BASE}{endpoint}", headers=headers, params=params) as r:
            return await r.json()

# =========================
# DIARY
# =========================

FILE = "Football_Model_Rules_and_Diary.xlsx"

def init_diary():
    if os.path.exists(FILE):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Diary"
    ws.append(["Date", "Match", "Market", "Odds", "Stake"])
    wb.save(FILE)

def add_diary(match, market, odds):
    wb = load_workbook(FILE)
    ws = wb.active
    ws.append([datetime.now().strftime("%Y-%m-%d %H:%M"), match, market, odds, STAKE])
    wb.save(FILE)

# =========================
# BOT
# =========================

bot = Bot(BOT_TOKEN)
dp = Dispatcher()

@dp.message(Command("start"))
async def start(msg: Message):
    if msg.from_user.id not in ADMIN_IDS:
        return
    await msg.answer("✅ Бот запущен и работает")

async def prematch_scanner():
    while True:
        data = await api_get("/fixtures", {"live": "all"})
        for f in data.get("response", []):
            home = f["teams"]["home"]["name"]
            away = f["teams"]["away"]["name"]
            text = f"🔴 LIVE\n{home} — {away}"
            await bot.send_message(CHANNEL_ID, text)
            add_diary(f"{home}-{away}", "LIVE CHECK", 1.85)
            break
        await asyncio.sleep(60)

async def main():
    init_diary()
    asyncio.create_task(prematch_scanner())
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
